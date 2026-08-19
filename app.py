from modules.community.reddit import (
    RedditCollector,
    run_community_intelligence,
)
from modules.research.literature import (
    ResearchCollector,
    run_research_intelligence,
)
from modules.search.intelligence import run_search_intelligence
from modules.strategy.funnel import run_funnel_strategy
from modules.brief.content_brief import run_content_brief
import io
import re
from modules.extractor import (
    process_competitors,
    extract_entities_for_document,
    serialize_competitor_entities,
    merge_entities_across_competitors,
    add_entity_co_occurrence,
    build_entity_dashboard,
    generate_section_topic_intelligence,
    generate_structure_intelligence,
)
import pandas as pd
import streamlit as st
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from modules.project.project_manager import (
    initialize,
    list_projects,
)

from modules.serp.serper import get_serp
from modules.scraping.collector import collect_competitors


st.set_page_config(
    page_title="SEO Intelligence Platform",
    page_icon="📈",
    layout="wide"
)

# Streamlit reserves a large top margin by default for its floating
# toolbar; trim it (main area + sidebar) and tighten the gap under the
# title so the page starts using screen space right away.
_LAYOUT_CSS = """
<style>
    .block-container {
        /* Must clear Streamlit's fixed top toolbar (~3.75rem tall) or it
           overlaps and clips the title/first widgets underneath it. */
        padding-top: 4rem !important;
        padding-bottom: 1.5rem !important;
    }
    h1 {
        margin-bottom: 0.25rem !important;
        padding-bottom: 0.25rem !important;
    }
</style>
"""
# The sidebar itself (API Keys / Insight Modules / Reddit) is left at
# Streamlit's default spacing, box size, and collapse arrow - custom
# compaction there caused more problems (clipped icons, overlap) than it
# solved. "Insight Modules" sits between API Keys and Reddit because of
# the order those st.sidebar.header()/checkbox() calls appear in the code
# below, not because of any CSS.

# st.html() (Streamlit >=1.41) injects raw HTML/CSS with no sanitization.
# Older Streamlit versions can silently strip <style> tags out of
# st.markdown(..., unsafe_allow_html=True) - no error, the CSS just never
# applies - which is why prior spacing tweaks here had zero visible effect
# no matter how large the values were. st.html sidesteps that entirely;
# markdown is kept only as a fallback for very old Streamlit installs.
if hasattr(st, "html"):
    st.html(_LAYOUT_CSS)
else:
    st.markdown(_LAYOUT_CSS, unsafe_allow_html=True)

initialize()

st.title("SEO Intelligence Platform")


def render_table(rows, empty_message):
    if rows:
        st.dataframe(
            rows,
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info(empty_message)


# Matches the Content Brief's own top-level section headings (A. SEARCH
# INTELLIGENCE ... I. SOURCES), tolerating the same markdown decoration
# (#, ##, **) the model sometimes adds around headings elsewhere.
_BRIEF_SECTION_HEADING_RE = re.compile(
    r"^[ \t]*(?:#{1,4}[ \t]*)?\*{0,2}[ \t]*"
    r"([A-I])\.[ \t]*\*{0,2}[ \t]*(.+?)\*{0,2}[ \t]*$"
)

# Same color as the app's primary (type="primary") buttons.
_BRAND_ACCENT_COLOR = "#FF4B4B"


def _md_safe(text):
    """Escape characters Streamlit's markdown renderer treats as special
    (notably '$', which it reads as LaTeX/math delimiters) so raw scraped
    text like dollar amounts displays as plain text instead of a formula."""
    return str(text or "").replace("$", "\\$")


def _colorize_brief_section_headings(brief_text, color=_BRAND_ACCENT_COLOR):
    """Bold + color the A./B./C.-style section headings in the Content
    Brief so they stand out, matching the primary button color. Everything
    else in the brief passes through untouched."""
    out_lines = []
    for line in (brief_text or "").split("\n"):
        match = _BRIEF_SECTION_HEADING_RE.match(line)
        if match:
            letter, label = match.group(1), match.group(2).strip()
            out_lines.append(
                f'<span style="color:{color}; font-weight:700;">'
                f"{letter}. {label}</span>"
            )
        else:
            out_lines.append(line)
    return "\n".join(out_lines)


def run_engine(competitors, do_entities=True, do_topics=True):
    """
    Stage-aware wrapper over the extractor. Competitor processing and
    structure always run (needed for the outline). Entity extraction and
    topic generation run only when enabled, so unticked stages are never
    computed. Graceful degradation:
    - entities off  -> topics + competitors + structure still available
    - topics off    -> entities + competitors + structure still available
    - both off      -> competitors + structure (SERP + scraped content) only
    """
    processed, docs = process_competitors(competitors or [])

    entities = []
    if do_entities:
        document_entities = []
        all_candidates = []
        for index, (competitor, doc) in enumerate(zip(processed, docs)):
            candidates = extract_entities_for_document(
                doc=doc,
                text=competitor.get("text", ""),
                structure=competitor.get("structure", {}),
                competitor=competitor,
                competitor_index=index,
            )
            competitor["entities"] = serialize_competitor_entities(candidates)
            document_entities.append(candidates)
            all_candidates.extend(candidates)
        aggregates = merge_entities_across_competitors(all_candidates)
        add_entity_co_occurrence(aggregates, document_entities)
        entities = build_entity_dashboard(aggregates, len(processed))

    topics = []
    if do_topics:
        topics = generate_section_topic_intelligence(processed, entities)

    structure = generate_structure_intelligence(processed)

    return {
        "competitors": processed,
        "entities": entities,
        "topics": topics,
        "structure": structure,
    }


def _safe_sheet_name(name, used):
    # Excel sheet names: max 31 chars, no : \ / ? * [ ], must be unique.
    name = re.sub(r"[:\\/?*\[\]]", " ", str(name)).strip()[:31] or "Sheet"
    base = name
    counter = 1
    while name in used:
        suffix = f"_{counter}"
        name = base[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(name)
    return name


def _cell(v):
    # Excel cells can't hold lists/dicts and cap at 32,767 chars.
    if isinstance(v, (list, tuple)):
        v = "; ".join(map(str, v))
    elif isinstance(v, dict):
        v = str(v)
    if isinstance(v, str) and len(v) > 32000:
        v = v[:32000] + " ...[truncated]"
    return v


# Colour is deliberately restrained: only the title band is filled. The
# header row and data rows stay white and are distinguished by bold navy
# text and a thin border instead of a background fill, so a wide table
# doesn't turn into a wall of blue.
_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
_HEADER_FONT = Font(bold=True, color="1F4E78")
_THIN_SIDE = Side(style="thin", color="D9DEE3")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_HEADER_BOTTOM_SIDE = Side(style="thin", color="1F4E78")
_HEADER_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_HEADER_BOTTOM_SIDE
)
_NOWRAP_DATA = Alignment(horizontal="left", vertical="center", wrap_text=False)
_WRAP_DATA = Alignment(horizontal="left", vertical="top", wrap_text=True)
_WRAP_HEADER = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _rows_to_table(rows):
    """
    Normalises a list of dicts into (columns, row_values), preserving
    first-seen column order even if some rows are missing keys others
    have. Returns ([], []) for empty input so the caller can skip it.
    """
    rows = [r for r in (rows or []) if r]
    columns = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)
    data = [[_cell(row.get(col, "")) for col in columns] for row in rows]
    return columns, data


def _write_table_block(ws, start_row, title, rows, col_width=32, max_col_width=60, wrap_data=False):
    """
    Writes ONE self-contained, titled, coloured table starting at
    start_row: a coloured title band, then its own header row with only
    that table's columns, then its data rows (never unioned with any
    other table's columns), then a blank spacer row. Returns the next
    free row. Writes nothing and returns start_row unchanged if there is
    no data, so empty tables don't leave a stray title band behind.

    wrap_data=False (default) keeps rows compact/single-line, which is
    what a data table (Topics, Entities, Reddit, etc.) needs - long
    values (e.g. a Reddit permalink) are simply clipped to the column
    width rather than blowing up the row height. Pass wrap_data=True only
    for prose content (the Content Brief sheet) where wrapping is what
    makes it readable.
    """
    columns, data = _rows_to_table(rows)
    if not columns:
        return start_row

    n_cols = max(len(columns), 1)

    title_cell = ws.cell(row=start_row, column=1, value=title)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=n_cols)
    title_cell.font = _TITLE_FONT
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 20

    header_row = start_row + 1
    for c_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=str(name))
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP_HEADER
        cell.border = _HEADER_BORDER

    data_alignment = _WRAP_DATA if wrap_data else _NOWRAP_DATA
    for r_offset, values in enumerate(data):
        r_idx = header_row + 1 + r_offset
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = data_alignment
            cell.border = _THIN_BORDER

    for c_idx, name in enumerate(columns, start=1):
        longest = max([len(str(name))] + [len(str(v)) for v in [row[c_idx - 1] for row in data]])
        col_letter = ws.cell(row=header_row, column=c_idx).column_letter
        current = ws.column_dimensions[col_letter].width or 0
        ws.column_dimensions[col_letter].width = max(current, min(max(longest + 2, 10), max_col_width))

    return header_row + 1 + len(data) + 1  # +1 blank spacer row


def _write_sheet(wb, sheet_name, blocks, max_col_width=60):
    """blocks: list of (title, rows) tuples, written one below the other."""
    ws = wb.create_sheet(sheet_name)
    row = 1
    wrote_any = False
    for title, rows in blocks:
        next_row = _write_table_block(ws, row, title, rows, max_col_width=max_col_width)
        if next_row != row:
            wrote_any = True
        row = next_row
    if not wrote_any:
        ws.cell(row=1, column=1, value="No data for this run.")
    ws.freeze_panes = "A2"
    return ws


_HEADER_TOKEN_RE = re.compile(
    r"(?<=\S)("
    r"#{1,4}\s"                    # markdown headers run into prior text
    r"|\*\*[A-Z][^*\n]{0,80}:\*\*"  # a **Bold Label:** run into prior text
    r"|H[1-4]\s*:"                  # H1:/H2:/H3:/H4: run into prior text
    r"|[A-I]\.\s+[A-Z][a-z]"        # 'A. Search Intelligence' style markers
    r")"
)


def _brief_to_rows(brief_text):
    """
    One Excel row per logical line of the brief. The model sometimes emits
    a header immediately followed by a bold label with no newline between
    them (e.g. '# A. SEARCH INTELLIGENCE**Primary Intent:** ...'), which
    would otherwise land in Excel as one unreadable run-on row. Force a
    newline before those structural markers first, then split on '\\n'.
    """
    if not brief_text:
        return []
    normalized = _HEADER_TOKEN_RE.sub(r"\n\1", brief_text)
    return [{"line": ln} for ln in normalized.split("\n") if ln.strip()]


def build_insights_xlsx(
    organic_results, competitors, topics, entities,
    search_intel, community, research, strategy, brief,
):
    """
    One sheet per tab. Each sub-table on a sheet gets its own coloured
    title band, its own header row, and its own data rows, stacked one
    below the other with a blank spacer row - never unioned into a single
    table with a 'section' discriminator column.
    """
    search_intel = search_intel or {}
    community = community or {}
    research = research or {}
    recommendations = (strategy or {}).get("recommendations", {}) or {}

    serp_rows = [
        {"title": r.get("title", ""), "link": r.get("link", ""),
         "snippet": r.get("snippet", "")}
        for r in (organic_results or [])
    ]
    competitor_rows = [
        {"position": c.get("position"), "title": c.get("title"),
         "url": c.get("url"), "words": len((c.get("text") or "").split()),
         "credits": c.get("credits")}
        for c in (competitors or [])
    ]
    rec_rows_mofu = [{"stage": "MOFU", **item} for item in recommendations.get("MOFU", []) or []]
    rec_rows_bofu = [{"stage": "BOFU", **item} for item in recommendations.get("BOFU", []) or []]
    rec_rows_faq = [{"stage": "FAQ", **item} for item in recommendations.get("FAQs", []) or []]
    if strategy and strategy.get("error"):
        rec_rows_mofu = [{"error": strategy["error"]}]

    if brief and brief.get("brief"):
        brief_rows = _brief_to_rows(brief["brief"])
    elif brief and brief.get("error"):
        brief_rows = [{"line": "Content brief failed: " + brief["error"]}]
    else:
        brief_rows = []
    stats = community.get("statistics", {})

    tabs = {
        "SERP": [("Organic Results", serp_rows)],
        "Search": [
            ("Signal Matrix", search_intel.get("signal_matrix", [])),
            ("Question Intelligence", search_intel.get("question_intelligence", [])),
            ("PAA", search_intel.get("paa", [])),
            ("Related Searches", search_intel.get("related_searches", [])),
            ("Autosuggest", search_intel.get("autosuggest", [])),
            ("FAQs", search_intel.get("faqs", [])),
        ],
        "Competitors": [("Competitors", competitor_rows)],
        "Topics": [("Topics", topics)],
        "Entities": [("Entities", entities)],
        "Reddit": [
            ("Statistics", [stats] if stats else []),
            ("Questions", community.get("questions", [])),
            ("Pain Points", community.get("pain_points", [])),
            ("Recommendations", community.get("recommendations", [])),
            ("Features", community.get("features", [])),
            ("Decision Factors", community.get("decision_factors", [])),
            ("Vocabulary", community.get("vocabulary", [])),
            ("Mistakes", community.get("mistakes", [])),
            ("Myths", community.get("myths", [])),
            ("Experiences", community.get("experiences", [])),
            ("Gaps", community.get("gaps", [])),
            ("Brands", community.get("brands", [])),
        ],
        "Research": [
            ("Papers", research.get("papers", [])),
            ("Statistics", research.get("data_points", [])),
            ("Information Gain", research.get("information_gain", [])),
            ("Concepts", research.get("concepts", [])),
        ],
        "Recommendations": [
            ("MOFU Ideas", rec_rows_mofu),
            ("BOFU Ideas", rec_rows_bofu),
            ("FAQs", rec_rows_faq),
        ],
    }

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for tab_name, blocks in tabs.items():
        sheet = _safe_sheet_name(tab_name, used)
        _write_sheet(wb, sheet, blocks)

    # Content Brief gets its own wide, wrapped column since it's prose,
    # not a data table.
    brief_sheet = _safe_sheet_name("Content Brief", used)
    ws = wb.create_sheet(brief_sheet)
    _write_table_block(ws, 1, "Content Brief", brief_rows, max_col_width=120, wrap_data=True)
    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

# -----------------------------
# Sidebar
# -----------------------------

st.sidebar.header("API Keys")

openai_key = st.sidebar.text_input(
    "OpenAI API Key",
    type="password"
)

serper_key = st.sidebar.text_input(
    "Serper API Key",
    type="password"
)

client = None

if openai_key:
    client = OpenAI(api_key=openai_key)

# Modules to run sits right after API Keys (not after Projects/Reddit
# credentials) so all six toggles are visible on the first screen without
# scrolling the sidebar.
st.sidebar.header("Insight Modules")

enable_serp_insights = st.sidebar.checkbox("SERP insights", value=True)
enable_reddit = st.sidebar.checkbox("Reddit insights", value=True)
enable_research = st.sidebar.checkbox("Research insights", value=True)
enable_topics = st.sidebar.checkbox("Topic generation", value=True)
enable_entities = st.sidebar.checkbox("Entity extraction", value=True)
enable_outline = st.sidebar.checkbox("Outline generation (Content Brief)", value=True)

st.sidebar.header("Reddit")

projects = list_projects()

if projects:
    st.sidebar.selectbox(
        "Existing Projects",
        projects
    )

reddit_client_id = st.sidebar.text_input(
    "Reddit Client ID",
    type="password"
)

reddit_client_secret = st.sidebar.text_input(
    "Reddit Client Secret",
    type="password"
)

reddit_user_agent = st.sidebar.text_input(
    "Reddit User Agent"
)

# -----------------------------
# Inputs
# -----------------------------

col1, col2, col3, col4 = st.columns(4)

with col1:
    keyword = st.text_input(
        "Keyword"
    )

with col2:
    country = st.selectbox(
        "Country",
        {
            "Afghanistan": "af",
            "Albania": "al",
            "Algeria": "dz",
            "American Samoa": "as",
            "Andorra": "ad",
            "Angola": "ao",
            "Anguilla": "ai",
            "Antarctica": "aq",
            "Antigua and Barbuda": "ag",
            "Argentina": "ar",
            "Armenia": "am",
            "Aruba": "aw",
            "Australia": "au",
            "Austria": "at",
            "Azerbaijan": "az",
            "Bahamas": "bs",
            "Bahrain": "bh",
            "Bangladesh": "bd",
            "Barbados": "bb",
            "Belarus": "by",
            "Belgium": "be",
            "Belize": "bz",
            "Benin": "bj",
            "Bermuda": "bm",
            "Bhutan": "bt",
            "Bolivia": "bo",
            "Bonaire, Sint Eustatius and Saba": "bq",
            "Bosnia and Herzegovina": "ba",
            "Botswana": "bw",
            "Bouvet Island": "bv",
            "Brazil": "br",
            "British Indian Ocean Territory": "io",
            "Brunei": "bn",
            "Bulgaria": "bg",
            "Burkina Faso": "bf",
            "Burundi": "bi",
            "Cabo Verde": "cv",
            "Cambodia": "kh",
            "Cameroon": "cm",
            "Canada": "ca",
            "Cayman Islands": "ky",
            "Central African Republic": "cf",
            "Chad": "td",
            "Chile": "cl",
            "China": "cn",
            "Christmas Island": "cx",
            "Cocos (Keeling) Islands": "cc",
            "Colombia": "co",
            "Comoros": "km",
            "Democratic Republic of the Congo": "cd",
            "Congo": "cg",
            "Cook Islands": "ck",
            "Costa Rica": "cr",
            "Croatia": "hr",
            "Cuba": "cu",
            "Curaçao": "cw",
            "Cyprus": "cy",
            "Czechia": "cz",
            "Côte d'Ivoire": "ci",
            "Denmark": "dk",
            "Djibouti": "dj",
            "Dominica": "dm",
            "Dominican Republic": "do",
            "Ecuador": "ec",
            "Egypt": "eg",
            "El Salvador": "sv",
            "Equatorial Guinea": "gq",
            "Eritrea": "er",
            "Estonia": "ee",
            "Eswatini": "sz",
            "Ethiopia": "et",
            "Falkland Islands (Malvinas)": "fk",
            "Faroe Islands": "fo",
            "Fiji": "fj",
            "Finland": "fi",
            "France": "fr",
            "French Guiana": "gf",
            "French Polynesia": "pf",
            "French Southern Territories": "tf",
            "Gabon": "ga",
            "Gambia": "gm",
            "Georgia": "ge",
            "Germany": "de",
            "Ghana": "gh",
            "Gibraltar": "gi",
            "Greece": "gr",
            "Greenland": "gl",
            "Grenada": "gd",
            "Guadeloupe": "gp",
            "Guam": "gu",
            "Guatemala": "gt",
            "Guernsey": "gg",
            "Guinea": "gn",
            "Guinea-Bissau": "gw",
            "Guyana": "gy",
            "Haiti": "ht",
            "Heard Island and McDonald Islands": "hm",
            "Vatican City": "va",
            "Honduras": "hn",
            "Hong Kong": "hk",
            "Hungary": "hu",
            "Iceland": "is",
            "India": "in",
            "Indonesia": "id",
            "Iran": "ir",
            "Iraq": "iq",
            "Ireland": "ie",
            "Isle of Man": "im",
            "Israel": "il",
            "Italy": "it",
            "Jamaica": "jm",
            "Japan": "jp",
            "Jersey": "je",
            "Jordan": "jo",
            "Kazakhstan": "kz",
            "Kenya": "ke",
            "Kiribati": "ki",
            "North Korea": "kp",
            "South Korea": "kr",
            "Kuwait": "kw",
            "Kyrgyzstan": "kg",
            "Laos": "la",
            "Latvia": "lv",
            "Lebanon": "lb",
            "Lesotho": "ls",
            "Liberia": "lr",
            "Libya": "ly",
            "Liechtenstein": "li",
            "Lithuania": "lt",
            "Luxembourg": "lu",
            "Macao": "mo",
            "Madagascar": "mg",
            "Malawi": "mw",
            "Malaysia": "my",
            "Maldives": "mv",
            "Mali": "ml",
            "Malta": "mt",
            "Marshall Islands": "mh",
            "Martinique": "mq",
            "Mauritania": "mr",
            "Mauritius": "mu",
            "Mayotte": "yt",
            "Mexico": "mx",
            "Micronesia": "fm",
            "Moldova": "md",
            "Monaco": "mc",
            "Mongolia": "mn",
            "Montenegro": "me",
            "Montserrat": "ms",
            "Morocco": "ma",
            "Mozambique": "mz",
            "Myanmar": "mm",
            "Namibia": "na",
            "Nauru": "nr",
            "Nepal": "np",
            "Netherlands": "nl",
            "New Caledonia": "nc",
            "New Zealand": "nz",
            "Nicaragua": "ni",
            "Niger": "ne",
            "Nigeria": "ng",
            "Niue": "nu",
            "Norfolk Island": "nf",
            "North Macedonia": "mk",
            "Northern Mariana Islands": "mp",
            "Norway": "no",
            "Oman": "om",
            "Pakistan": "pk",
            "Palau": "pw",
            "Palestine": "ps",
            "Panama": "pa",
            "Papua New Guinea": "pg",
            "Paraguay": "py",
            "Peru": "pe",
            "Philippines": "ph",
            "Pitcairn": "pn",
            "Poland": "pl",
            "Portugal": "pt",
            "Puerto Rico": "pr",
            "Qatar": "qa",
            "Romania": "ro",
            "Russia": "ru",
            "Rwanda": "rw",
            "Réunion": "re",
            "Saint Barthélemy": "bl",
            "Saint Helena, Ascension and Tristan da Cunha": "sh",
            "Saint Kitts and Nevis": "kn",
            "Saint Lucia": "lc",
            "Saint Martin (French part)": "mf",
            "Saint Pierre and Miquelon": "pm",
            "Saint Vincent and the Grenadines": "vc",
            "Samoa": "ws",
            "San Marino": "sm",
            "Sao Tome and Principe": "st",
            "Saudi Arabia": "sa",
            "Senegal": "sn",
            "Serbia": "rs",
            "Seychelles": "sc",
            "Sierra Leone": "sl",
            "Singapore": "sg",
            "Sint Maarten (Dutch part)": "sx",
            "Slovakia": "sk",
            "Slovenia": "si",
            "Solomon Islands": "sb",
            "Somalia": "so",
            "South Africa": "za",
            "South Georgia and the South Sandwich Islands": "gs",
            "South Sudan": "ss",
            "Spain": "es",
            "Sri Lanka": "lk",
            "Sudan": "sd",
            "Suriname": "sr",
            "Svalbard and Jan Mayen": "sj",
            "Sweden": "se",
            "Switzerland": "ch",
            "Syria": "sy",
            "Taiwan": "tw",
            "Tajikistan": "tj",
            "Tanzania": "tz",
            "Thailand": "th",
            "Timor-Leste": "tl",
            "Togo": "tg",
            "Tokelau": "tk",
            "Tonga": "to",
            "Trinidad and Tobago": "tt",
            "Tunisia": "tn",
            "Turkmenistan": "tm",
            "Turks and Caicos Islands": "tc",
            "Tuvalu": "tv",
            "Türkiye": "tr",
            "Uganda": "ug",
            "Ukraine": "ua",
            "United Arab Emirates": "ae",
            "United Kingdom": "gb",
            "United States Minor Outlying Islands": "um",
            "United States": "us",
            "Uruguay": "uy",
            "Uzbekistan": "uz",
            "Vanuatu": "vu",
            "Venezuela": "ve",
            "Vietnam": "vn",
            "British Virgin Islands": "vg",
            "U.S. Virgin Islands": "vi",
            "Wallis and Futuna": "wf",
            "Western Sahara": "eh",
            "Yemen": "ye",
            "Zambia": "zm",
            "Zimbabwe": "zw",
            "Åland Islands": "ax",
        }
    )

with col3:
    language = st.selectbox(
        "Language",
        {
            "Afar": "aa",
            "Abkhazian": "ab",
            "Avestan": "ae",
            "Afrikaans": "af",
            "Akan": "ak",
            "Amharic": "am",
            "Aragonese": "an",
            "Arabic": "ar",
            "Assamese": "as",
            "Avaric": "av",
            "Aymara": "ay",
            "Azeri": "az",
            "Bashkir": "ba",
            "Belarusian": "be",
            "Bulgarian": "bg",
            "Bihari": "bh",
            "Bislama": "bi",
            "Bambara": "bm",
            "Bengali": "bn",
            "Tibetan": "bo",
            "Breton": "br",
            "Bosnian": "bs",
            "Catalan": "ca",
            "Chechen": "ce",
            "Chamorro": "ch",
            "Corsican": "co",
            "Cree": "cr",
            "Czech": "cs",
            "Church Slavonic": "cu",
            "Chuvash": "cv",
            "Welsh": "cy",
            "Danish": "da",
            "German": "de",
            "Divehi": "dv",
            "Bhutani": "dz",
            "Ewe": "ee",
            "Greek": "el",
            "English": "en",
            "Esperanto": "eo",
            "Spanish": "es",
            "Estonian": "et",
            "Basque": "eu",
            "Farsi": "fa",
            "Fulah": "ff",
            "Finnish": "fi",
            "Fiji": "fj",
            "Faroese": "fo",
            "French": "fr",
            "Frisian": "fy",
            "Irish": "ga",
            "Gaelic": "gd",
            "Galician": "gl",
            "Guarani": "gn",
            "Gujarati": "gu",
            "Manx": "gv",
            "Hausa": "ha",
            "Hebrew": "he",
            "Hindi": "hi",
            "Hiri Motu": "ho",
            "Croatian": "hr",
            "Haitian": "ht",
            "Hungarian": "hu",
            "Armenian": "hy",
            "Herero": "hz",
            "Interlingua": "ia",
            "Indonesian": "id",
            "Interlingue": "ie",
            "Igbo": "ig",
            "Sichuan Yi": "ii",
            "Inupiak": "ik",
            "Ido": "io",
            "Icelandic": "is",
            "Italian": "it",
            "Inuktitut": "iu",
            "Japanese": "ja",
            "Yiddish": "yi",
            "Javanese": "jv",
            "Georgian": "ka",
            "Kongo": "kg",
            "Kikuyu": "ki",
            "Kuanyama": "kj",
            "Kazakh": "kk",
            "Greenlandic": "kl",
            "Cambodian": "km",
            "Kannada": "kn",
            "Korean": "ko",
            "Konkani": "kok",
            "Kanuri": "kr",
            "Kashmiri": "ks",
            "Kurdish": "ku",
            "Komi": "kv",
            "Cornish": "kw",
            "Kirghiz": "ky",
            "Kyrgyz": "kz",
            "Latin": "la",
            "Luxembourgish": "lb",
            "Ganda": "lg",
            "Limburgan": "li",
            "Lingala": "ln",
            "Laothian": "lo",
            "Slovenian": "sl",
            "Lithuanian": "lt",
            "Luba-Katanga": "lu",
            "Latvian": "lv",
            "Malagasy": "mg",
            "Marshallese": "mh",
            "Maori": "mi",
            "FYRO Macedonian": "mk",
            "Malayalam": "ml",
            "Mongolian": "mn",
            "Moldavian": "mo",
            "Marathi": "mr",
            "Malay": "ms",
            "Maltese": "mt",
            "Burmese": "my",
            "Nauru": "na",
            "Norwegian (Bokmal)": "nb",
            "North Ndebele": "nd",
            "Nepali": "ne",
            "Ndonga": "ng",
            "Dutch": "nl",
            "Norwegian (Nynorsk)": "nn",
            "Norwegian": "no",
            "South Ndebele": "nr",
            "Northern Sotho": "ns",
            "Navajo": "nv",
            "Chichewa": "ny",
            "Occitan": "oc",
            "Ojibwa": "oj",
            "(Afan)/Oromoor/Oriya": "om",
            "Oriya": "or",
            "Ossetian": "os",
            "Punjabi": "pa",
            "Pali": "pi",
            "Polish": "pl",
            "Pashto/Pushto": "ps",
            "Portuguese": "pt",
            "Quechua": "qu",
            "Rhaeto-Romanic": "rm",
            "Kirundi": "rn",
            "Romanian": "ro",
            "Russian": "ru",
            "Kinyarwanda": "rw",
            "Sanskrit": "sa",
            "Sorbian": "sb",
            "Sardinian": "sc",
            "Sindhi": "sd",
            "Sami": "se",
            "Sangro": "sg",
            "Serbo-Croatian": "sh",
            "Singhalese": "si",
            "Slovak": "sk",
            "Samoan": "sm",
            "Shona": "sn",
            "Somali": "so",
            "Albanian": "sq",
            "Serbian": "sr",
            "Siswati": "ss",
            "Sesotho": "st",
            "Sundanese": "su",
            "Swedish": "sv",
            "Swahili": "sw",
            "Sutu": "sx",
            "Syriac": "syr",
            "Tamil": "ta",
            "Telugu": "te",
            "Tajik": "tg",
            "Thai": "th",
            "Tigrinya": "ti",
            "Turkmen": "tk",
            "Tagalog": "tl",
            "Tswana": "tn",
            "Tonga": "to",
            "Turkish": "tr",
            "Tsonga": "ts",
            "Tatar": "tt",
            "Twi": "tw",
            "Tahitian": "ty",
            "Uighur": "ug",
            "Ukrainian": "uk",
            "Urdu": "ur",
            "Uzbek": "uz",
            "Venda": "ve",
            "Vietnamese": "vi",
            "Volapuk": "vo",
            "Walloon": "wa",
            "Wolof": "wo",
            "Xhosa": "xh",
            "Yoruba": "yo",
            "Zhuang": "za",
            "Chinese": "zh",
            "Zulu": "zu",
        }
    )

with col4:
    num_results = st.slider(
        "Competitors",
        5,
        20,
        10
    )

# -----------------------------
# Run
# -----------------------------

# Run (left) and Download (right) share one row, both coloured via
# type="primary" instead of Streamlit's default plain grey buttons. The
# download button itself is only ever available after a run has produced
# results, but its slot in the right column is reserved here with
# st.empty() and filled in further down once insights_xlsx exists --
# that's what keeps it visually pinned to the right of Run rather than
# stacked underneath it.
run_col, download_col = st.columns([1, 1])
with run_col:
    run_clicked = st.button(
        "Run Competitor Intelligence",
        type="primary",
        use_container_width=True,
    )
with download_col:
    download_slot = st.empty()

if run_clicked:

    if not serper_key:
        st.error("Please enter Serper API Key.")
        st.stop()

    if not keyword:
        st.error("Please enter a keyword.")
        st.stop()

    project_name = keyword.lower().strip().replace(" ", "-")

    with st.spinner("Searching Google..."):

        organic_results, serp_summary = get_serp(
            keyword=keyword,
            serper_key=serper_key,
            country=country,
            language=language,
            num_results=num_results
        )

    with st.spinner("Scraping competitors..."):

        competitors = collect_competitors(
            keyword=keyword,
            serper_key=serper_key,
            project_name=project_name,
            country=country,
            language=language,
            num_results=num_results,
            organic_results=organic_results
        )

    stage_label = "Processing competitors"
    if enable_entities:
        stage_label = "Extracting entities"
    elif enable_topics:
        stage_label = "Generating topics"
    with st.spinner(stage_label + "..."):

        results = run_engine(
            competitors,
            do_entities=enable_entities,
            do_topics=enable_topics,
        )

        competitors = results["competitors"]
        entities = results["entities"]
        topics = results["topics"]

    reddit_threads = []
    community = {}

    if enable_reddit:
        if reddit_client_id and reddit_client_secret and reddit_user_agent:
            with st.spinner("Fetching Reddit discussions..."):
                collector = RedditCollector(
                    client_id=reddit_client_id,
                    client_secret=reddit_client_secret,
                    user_agent=reddit_user_agent,
                )
                reddit_threads = collector.search_via_google(
                    keyword=keyword,
                    serper_key=serper_key,
                    limit=10,
                    country=country,
                    language=language,
                )
            if reddit_threads:
                with st.spinner("Processing Reddit intelligence..."):
                    community = run_community_intelligence(
                        reddit_threads,
                        competitor_entities=entities,
                        competitor_topics=topics,
                    )

    research = {}
    research_papers = []

    if enable_research:
        with st.spinner("Scanning research literature..."):
            research_papers = ResearchCollector().search(
                keyword=keyword,
                limit=25,
            )
            if research_papers:
                research = run_research_intelligence(
                    research_papers,
                    competitor_entities=entities,
                    competitor_topics=topics,
                )

    search_intel = {}

    if enable_serp_insights:
        with st.spinner("Analyzing search intent..."):
            search_intel = run_search_intelligence(
                keyword=keyword,
                serper_key=serper_key,
                country=country,
                language=language,
                competitors=competitors,
                community=community,
            )

    strategy = {}

    if client is not None:
        with st.spinner("Generating funnel content strategy..."):
            strategy = run_funnel_strategy(
                client,
                keyword,
                topics=topics,
                entities=entities,
                community=community,
                search_intel=search_intel,
                competitors=competitors,
            )

    brief = {}

    if enable_outline and client is not None:
        with st.spinner("Researching all tabs and drafting content brief..."):
            brief = run_content_brief(
                client,
                keyword,
                bundle={
                    "keyword": keyword,
                    "topics": topics,
                    "entities": entities,
                    "search_intel": search_intel,
                    "community": community,
                    "reddit_threads": reddit_threads,
                    "research": research,
                    "research_papers": research_papers,
                    "competitors": competitors,
                    "strategy": strategy,
                },
            )

    # Persist results across reruns. Streamlit reruns the ENTIRE script on
    # every widget interaction, including clicking the download buttons
    # below. Without this, organic_results/competitors/etc. are local
    # variables that only exist during the run where this button was
    # actually clicked -- on the very next rerun (e.g. clicking "Download
    # all insights") st.button(...) is False again, this whole block is
    # skipped, and every tab disappears. Stashing the results here lets the
    # rendering below survive any later rerun.
    st.session_state["run_results"] = {
        "organic_results": organic_results,
        "competitors": competitors,
        "topics": topics,
        "entities": entities,
        "reddit_threads": reddit_threads,
        "community": community,
        "research": research,
        "research_papers": research_papers,
        "search_intel": search_intel,
        "strategy": strategy,
        "brief": brief,
        "project_name": project_name,
    }

# Render from session_state (not gated on the button) so downloading a
# file -- which is itself a button click and triggers a full rerun -- does
# not wipe the tabs. This block runs on every rerun as long as a result
# set exists, whether that rerun was caused by "Run Competitor
# Intelligence", a download button, or a sidebar toggle.
if "run_results" in st.session_state:
    run = st.session_state["run_results"]
    organic_results = run["organic_results"]
    competitors = run["competitors"]
    topics = run["topics"]
    entities = run["entities"]
    reddit_threads = run["reddit_threads"]
    community = run["community"]
    research = run["research"]
    research_papers = run["research_papers"]
    search_intel = run["search_intel"]
    strategy = run["strategy"]
    brief = run["brief"]
    project_name = run["project_name"]

    try:
        insights_xlsx = build_insights_xlsx(
            organic_results, competitors, topics, entities,
            search_intel, community, research, strategy, brief,
        )
        with download_slot:
            st.download_button(
                "Download all insights (Excel)",
                data=insights_xlsx,
                file_name=f"{project_name}-insights.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
    except Exception as error:
        st.warning(f"Could not build the insights workbook: {error}")

    tabs = st.tabs([
        "SERP",
        "Search",
        "Competitors",
        "Topics",
        "Entities",
        "Reddit",
        "Research",
        "Recommendations",
        "Content Brief"
    ])

    # -----------------------------
    # SERP
    # -----------------------------

    with tabs[0]:

        st.subheader("Organic Results")

        if not organic_results:
            st.info("No organic results found.")
        else:
            for position, result in enumerate(organic_results, start=1):
                title = _md_safe(result.get("title", "")) or "(untitled)"
                link = result.get("link", "")
                snippet = _md_safe(result.get("snippet", ""))

                with st.container(border=True):
                    st.markdown(f"**{position}. {title}**")
                    if link:
                        st.caption(link)
                    if snippet:
                        st.write(snippet)

    # -----------------------------
    # Search Intelligence
    # -----------------------------

    with tabs[1]:

        st.subheader("Search Intelligence")

        if not enable_serp_insights:
            st.info("SERP insights disabled. Enable it in the sidebar.")
        else:
            st.write("### Signal Matrix")
            st.caption(
                "Questions ranked by opportunity: strong Google presence + "
                "Reddit demand + low competitor coverage rise to the top."
            )
            render_table(
                search_intel.get("signal_matrix", []),
                "No signal matrix computed."
            )

            st.divider()
            st.write("### Question Intelligence")
            st.caption("Merged and clustered across all sources below.")
            render_table(
                search_intel.get("question_intelligence", []),
                "No questions found."
            )

            st.divider()
            st.write("### People Also Ask")
            render_table(search_intel.get("paa", []), "No PAA results.")

            st.divider()
            st.write("### Related Searches")
            render_table(
                search_intel.get("related_searches", []),
                "No related searches."
            )

            st.divider()
            st.write("### Autosuggest")
            render_table(search_intel.get("autosuggest", []), "No autosuggest results.")

            st.divider()
            st.write("### Competitor FAQs")
            render_table(search_intel.get("faqs", []), "No FAQs extracted.")

    # -----------------------------
    # Competitors
    # -----------------------------

    with tabs[2]:

        st.subheader("Competitor Intelligence")

        for competitor in competitors:

            with st.expander(competitor["title"]):

                col1, col2 = st.columns(2)

                with col1:
                    st.write("**Position:**", competitor["position"])
                    st.write("**URL:**", competitor["url"])

                with col2:
                    st.write("**Credits:**", competitor["credits"])
                    st.write("**Words:**", len(competitor["text"].split()))

                st.write("### Metadata")
                st.json(competitor["metadata"])

                st.write("### Content")

                st.text_area(
                    f"content_{competitor['position']}",
                    competitor["text"],
                    height=350
                )

    # -----------------------------
    # Topics
    # -----------------------------

    with tabs[3]:

        st.subheader("Topic Intelligence")

        if topics:
            topics_df = pd.DataFrame(topics)
            topic_columns = [
                "topic",
                "coverage",
                "competitors_using",
                "importance",
                "heading_hits",
                "sources",
                "urls",
                "average_word_count",
                "related_phrases",
                "related_entities",
            ]
            topic_columns = [c for c in topic_columns if c in topics_df.columns]
            st.dataframe(
                topics_df[topic_columns],
                use_container_width=True,
                hide_index=True
            )
        elif not enable_topics:
            st.info("Topic generation disabled. Enable it in the sidebar.")
        else:
            st.info("No topics found.")

    # -----------------------------
    # Entities
    # -----------------------------

    with tabs[4]:

        st.subheader("Entity Intelligence")

        if entities:
            entities_df = pd.DataFrame(entities)
            entity_columns = [
                "entity",
                "type",
                "mentions",
                "competitors_using",
                "coverage",
                "average_mentions",
                "importance",
                "confidence",
                "extractor_sources",
            ]
            entity_columns = [c for c in entity_columns if c in entities_df.columns]
            st.dataframe(
                entities_df[entity_columns],
                use_container_width=True,
                hide_index=True
            )
        elif not enable_entities:
            st.info("Entity extraction disabled. Enable it in the sidebar.")
        else:
            st.info("No entities found.")

    # -----------------------------
    # Reddit
    # -----------------------------

    with tabs[5]:

        st.subheader("Reddit Intelligence")

        if not enable_reddit:
            st.info("Reddit insights disabled. Enable it in the sidebar.")
        elif not reddit_threads:
            st.info("No Reddit threads collected (check Reddit credentials).")
        else:
            # Overview
            stats = community.get("statistics", {})
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Threads analyzed", stats.get("threads", 0))
            m2.metric("Comments analyzed", stats.get("comments", 0))
            m3.metric("Subreddits", stats.get("subreddits", 0))
            m4.metric("Avg. upvotes", stats.get("average_upvotes", 0))
            m5.metric("Time span", stats.get("time_span", "-"))
            if stats.get("subreddit_names"):
                st.caption(
                    "Subreddits: "
                    + ", ".join(f"r/{name}" for name in stats["subreddit_names"])
                )

            st.divider()
            st.subheader("Questions")
            render_table(community.get("questions", []), "No questions found.")

            st.divider()
            st.subheader("Pain Points")
            render_table(community.get("pain_points", []), "No pain points found.")

            st.divider()
            st.subheader("Recommendations")
            render_table(
                community.get("recommendations", []),
                "No recommendations found."
            )

            st.divider()
            st.subheader("Brands & Products")
            render_table(community.get("brands", []), "No brands found.")

            st.divider()
            st.subheader("Features Users Care About")
            render_table(community.get("features", []), "No features found.")

            st.divider()
            st.subheader("Decision Factors")
            render_table(
                community.get("decision_factors", []),
                "No decision factors found."
            )

            st.divider()
            st.subheader("Community Terminology")
            render_table(community.get("vocabulary", []), "No terminology found.")

            st.divider()
            st.subheader("Common Mistakes")
            render_table(community.get("mistakes", []), "No mistakes found.")

            st.divider()
            st.subheader("Myths")
            render_table(community.get("myths", []), "No myths found.")

            st.divider()
            st.subheader("Real Experiences")
            render_table(community.get("experiences", []), "No experiences found.")

            st.divider()
            st.subheader("Content Opportunities")
            render_table(
                community.get("gaps", []),
                "No content opportunities computed."
            )

            st.divider()
            st.subheader("Threads")
            st.write(f"Threads found: {len(reddit_threads)}")
            for thread in reddit_threads:
                with st.expander(thread.title):
                    st.write(f"**Subreddit:** r/{thread.subreddit}")
                    st.write(f"**Score:** {thread.score}")
                    st.write(f"**Comments:** {thread.num_comments}")
                    st.write(thread.selftext)
                    st.markdown(thread.permalink)
                    st.subheader("Top Comments")
                    for comment in thread.comments[:10]:
                        st.write(comment.body)

    # -----------------------------
    # Research
    # -----------------------------

    with tabs[6]:

        st.subheader("Research Insights")

        if not enable_research:
            st.info("Research insights disabled. Enable it in the sidebar.")
        elif not research:
            st.info("No research literature found.")
        else:
            # Overview
            rstats = research.get("statistics", {})
            r1, r2, r3, r4, r5 = st.columns(5)
            r1.metric("Papers scanned", rstats.get("papers", 0))
            r2.metric("Reviews", rstats.get("reviews", 0))
            r3.metric("Journals", rstats.get("journals", 0))
            r4.metric("Avg. citations", rstats.get("average_citations", 0))
            r5.metric("Year span", rstats.get("year_span", "-"))

            st.divider()
            st.subheader("Papers")
            render_table(
                research.get("papers", []),
                "No papers found."
            )

            st.divider()
            st.subheader("Statistics")
            render_table(
                research.get("data_points", []),
                "No numeric findings extracted."
            )

            st.divider()
            st.subheader("Information Gain")
            st.caption(
                "Research concepts ranked by lowest competitor coverage "
                "and highest research support. These are the unique, "
                "citable angles competitors miss."
            )
            render_table(
                research.get("information_gain", []),
                "No information-gain concepts computed."
            )

            st.divider()
            st.subheader("Research Concepts")
            render_table(
                research.get("concepts", []),
                "No research concepts found."
            )

    # -----------------------------
    # Recommendations (funnel strategy)
    # -----------------------------

    with tabs[7]:

        st.subheader("Content Recommendations")

        if client is None:
            st.info("Enter an OpenAI API key in the sidebar to generate "
                    "funnel content recommendations.")
        elif strategy.get("error"):
            st.error("Recommendations failed: " + strategy["error"])
        elif not strategy:
            st.info("No recommendations generated.")
        else:
            stage = strategy.get("stage", "")
            rationale = strategy.get("stage_rationale", "")
            st.write(f"**Detected funnel stage:** {stage}")
            if rationale:
                st.caption(rationale)
            st.caption(
                "Educational, brand-neutral angles for the stages below "
                "this keyword in the funnel."
            )

            recommendations = strategy.get("recommendations", {}) or {}

            for stage_key in ("MOFU", "BOFU"):
                items = recommendations.get(stage_key, [])
                if items:
                    st.divider()
                    st.subheader(f"{stage_key} Content Ideas")
                    render_table(items, f"No {stage_key} recommendations.")

            faqs = recommendations.get("FAQs", [])
            if faqs:
                st.divider()
                st.subheader("FAQs")
                render_table(faqs, "No FAQ recommendations.")

    # -----------------------------
    # Content Brief (research agent)
    # -----------------------------

    with tabs[8]:

        st.subheader("Content Brief")

        if not enable_outline:
            st.info("Outline generation disabled. Enable it in the sidebar.")
        elif client is None:
            st.info("Enter an OpenAI API key in the sidebar to generate a "
                    "content brief.")
        elif brief.get("error"):
            st.error("Content brief failed: " + brief["error"])
            st.caption("If this mentions the model, your OpenAI key may not "
                       "have access to it. Change BRIEF_MODEL in "
                       "modules/brief/content_brief.py (e.g. to gpt-4o-mini).")
        elif not brief.get("brief"):
            st.info("No content brief generated.")
        else:
            st.markdown(
                _colorize_brief_section_headings(brief["brief"]),
                unsafe_allow_html=True,
            )

            with st.expander("All sources"):
                sources = brief.get("sources", {})
                st.write("**Competitors**")
                render_table(sources.get("competitors", []), "None.")
                st.write("**Reddit threads**")
                render_table(sources.get("reddit_threads", []), "None.")
                st.write("**Papers**")
                render_table(sources.get("papers", []), "None.")